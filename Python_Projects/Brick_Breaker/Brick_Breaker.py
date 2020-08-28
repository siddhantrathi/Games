import pygame
import time
import sys
import openpyxl as xl
sys.path.insert(1, '/home/aradhya/Desktop/pygame_display/')
from display import Display
pygame.init()

# loading the workbook
book = xl.load_workbook('highscore.xlsx')
sheet = book.active

#display
width = 800
height = 650
dis = Display(title='Brick Breaker', width=width, height=height)
screen = dis.window()
clock = pygame.time.Clock()

#color
black = (0, 0, 0)
white = (255, 255, 255)

# functions

def player_animation():
    global player_speed
    player.x += player_speed
    if player.x <= 0:

        player.x = 0

    if player.x >= width - 150:
        player.x = width - 150


def ball_animation():
    global ball_speed_x, ball_speed_y, score_value, highscore
    ball.y += ball_speed_y
    ball.x += ball_speed_x

    if ball.y >= height - 30:
        game_over()



    if ball.y <= 0:
        ball_speed_y *= -1

    if ball.x >= width - 30 or ball.x <= 0:
        ball_speed_x *= -1

    if ball.colliderect(player) and ball_speed_y > 0:
        if abs(ball.bottom - player.top) < 30:
            ball_speed_y *= -1
            score_value += 1

        if score_value > sheet.cell(row=1, column=1, ).value:
            sheet.cell(row=1, column=1).value = score_value

def visuals():
    screen.fill(white)
    pygame.draw.rect(screen, black, player)
    pygame.draw.ellipse(screen, black, ball)
    screen.blit(score, (0,0))
    screen.blit(highscore, (300, 0))

def game_over():
    time.sleep(1)
    book.save('highscore.xlsx')
    sys.exit()




# game rectangles
player = pygame.Rect(width / 2 - 75, height - 20, 150, 20)
ball = pygame.Rect(width / 2 - 30, height / 2 - 30, 30, 30)

# player
player_speed = 0

#ball
ball_speed_x = -1
ball_speed_y = -1

#Score
score_value = 0
font = pygame.font.SysFont('monospace', 40, bold=True)


while True:
    #book.save('highscore.xlsx')
    #
    score = font.render(f'Score: {score_value}', 1, (0, 0, 0))
    highscore = font.render(f'High Score: {sheet.cell(row=1, column=1, ).value}', 1, (0, 0, 0))
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            book.save('highscore.xlsx')
            pygame.quit
            sys.exit()

        if event.type == pygame.KEYDOWN:
            if event.key == pygame.K_RIGHT:
                player_speed = 1

            if event.key == pygame.K_LEFT:
                player_speed = -1

        if event.type == pygame.KEYUP:
            if event.key == pygame.K_RIGHT or event.key == pygame.K_LEFT:
                player_speed = 0


    player_animation()
    ball_animation()
    visuals()



    pygame.display.flip()
    clock.tick(600)
